from __future__ import annotations

import csv
import hashlib
import io
import json
import os
import re
import secrets
import unicodedata
from datetime import datetime, timedelta, timezone
from functools import wraps
from pathlib import Path

from dotenv import load_dotenv
from flask import Flask, Response, current_app, jsonify, redirect, render_template, request, send_from_directory
from flask_cors import CORS
from flask_sqlalchemy import SQLAlchemy
from openpyxl import Workbook, load_workbook
from reportlab.lib import colors
from reportlab.lib.pagesizes import landscape, letter
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
from sqlalchemy import or_
from werkzeug.exceptions import RequestEntityTooLarge
from werkzeug.security import check_password_hash, generate_password_hash


BASE_DIR = Path(__file__).resolve().parent
MOBILE_APP_DIR = BASE_DIR / "mobile_app"
DEFAULT_CORS_ORIGINS = "https://cisternas-api-wqac.onrender.com,null"
DEFAULT_UPLOAD_LIMIT_BYTES = 5 * 1024 * 1024
FORMULA_PREFIXES = ("=", "+", "-", "@")
FORMULA_LEADING_CHARS = ("\t", "\r", "\n")
INITIAL_USERS = [
    ("admin", "ADMIN_PASSWORD", "admin"),
    ("usuario", "USER_PASSWORD", "user"),
    ("cliente", "CLIENT_PASSWORD", "user"),
]
WEAK_INITIAL_PASSWORDS = {"admin123", "usuario123", "cliente123"}
load_dotenv(BASE_DIR / ".env")

db = SQLAlchemy()
LOGIN_ATTEMPTS: dict[str, list[datetime]] = {}

RECORD_EXPORT_HEADERS = [
    "ID",
    "Fecha UTC",
    "Conductor",
    "Placa",
    "Codigo funcionario",
    "EBAP",
    "Lectura inicial",
    "Lectura final",
    "Volumen",
    "Tipo empresa",
    "Empresa/Entidad",
    "Caracteristicas",
    "Registrado por",
]

IMPORT_ALIASES = {
    "timestamp": {"fechautc", "fecha", "timestamp", "timestampiso", "date"},
    "driver_name": {"conductor", "chofer", "driver", "drivername", "drivername"},
    "plate_number": {"placa", "plate", "platenumber", "platenumber"},
    "employee_code": {"codigofuncionario", "funcionario", "employeecode", "employeecode", "codigo"},
    "ebap": {"ebap"},
    "initial_reading": {"lecturainicial", "inicial", "initialreading", "initialreading"},
    "final_reading": {"lecturafinal", "final", "finalreading", "finalreading"},
    "load_volume": {"volumen", "volumencalculado", "loadvolume", "loadvolume"},
    "company_type": {"tipoempresa", "tipo", "companytype", "companytype"},
    "company_name": {"empresaentidad", "empresa", "entidad", "companyname", "companyname"},
    "characteristics": {"caracteristicas", "observaciones", "characteristics"},
    "registered_by": {"registradopor", "usuario", "registeredby", "registeredby"},
}


def utcnow() -> datetime:
    return datetime.now(timezone.utc)


def as_utc(value: datetime) -> datetime:
    if value.tzinfo is None:
        return value.replace(tzinfo=timezone.utc)
    return value.astimezone(timezone.utc)


def to_timestamp_ms(value: datetime) -> int:
    return int(as_utc(value).timestamp() * 1000)


def normalize_database_url(value: str) -> str:
    if value.startswith("postgres://"):
        return "postgresql://" + value[len("postgres://") :]
    return value


def configured_database_uri(app: Flask) -> str:
    database_url = os.getenv("DATABASE_URL", "").strip()
    if database_url:
        return normalize_database_url(database_url)
    return "sqlite:///" + str(Path(app.instance_path) / "cisternas.db")


class User(db.Model):
    __tablename__ = "users"

    username = db.Column(db.String(80), primary_key=True)
    password_hash = db.Column(db.String(255), nullable=False)
    role = db.Column(db.String(30), nullable=False, default="user")
    active = db.Column(db.Boolean, nullable=False, default=True)
    created_at = db.Column(db.DateTime(timezone=True), nullable=False, default=utcnow)


class Session(db.Model):
    __tablename__ = "sessions"

    id = db.Column(db.Integer, primary_key=True)
    token_hash = db.Column(db.String(64), nullable=False, unique=True, index=True)
    username = db.Column(db.String(80), db.ForeignKey("users.username"), nullable=False)
    expires_at = db.Column(db.DateTime(timezone=True), nullable=False)
    created_at = db.Column(db.DateTime(timezone=True), nullable=False, default=utcnow)
    last_seen_at = db.Column(db.DateTime(timezone=True), nullable=False, default=utcnow)

    user = db.relationship("User")


class WaterRecord(db.Model):
    __tablename__ = "registros_agua"

    id = db.Column(db.Integer, primary_key=True)
    timestamp = db.Column(db.DateTime(timezone=True), nullable=False, default=utcnow, index=True)
    driver_name = db.Column(db.String(160), nullable=False)
    plate_number = db.Column(db.String(60), nullable=False, index=True)
    employee_code = db.Column(db.String(80), nullable=False)
    ebap = db.Column(db.String(80), nullable=False, index=True)
    initial_reading = db.Column(db.Float, nullable=False)
    final_reading = db.Column(db.Float, nullable=False)
    load_volume = db.Column(db.Float, nullable=False)
    company_type = db.Column(db.String(120), nullable=False, index=True)
    company_name = db.Column(db.String(180), nullable=False, index=True)
    characteristics = db.Column(db.Text, nullable=True)
    registered_by = db.Column(db.String(80), db.ForeignKey("users.username"), nullable=False, index=True)
    created_at = db.Column(db.DateTime(timezone=True), nullable=False, default=utcnow)

    user = db.relationship("User")


class AuditLog(db.Model):
    __tablename__ = "audit_logs"

    id = db.Column(db.Integer, primary_key=True)
    timestamp = db.Column(db.DateTime(timezone=True), nullable=False, default=utcnow, index=True)
    actor = db.Column(db.String(80), nullable=True, index=True)
    action = db.Column(db.String(80), nullable=False, index=True)
    target_type = db.Column(db.String(80), nullable=True, index=True)
    target_id = db.Column(db.String(120), nullable=True, index=True)
    ip_address = db.Column(db.String(80), nullable=True)
    user_agent = db.Column(db.String(255), nullable=True)
    details = db.Column(db.Text, nullable=True)


def create_app(test_config: dict | None = None) -> Flask:
    app = Flask(__name__, instance_relative_config=True)
    Path(app.instance_path).mkdir(parents=True, exist_ok=True)

    app.config.update(
        SECRET_KEY=os.getenv("SECRET_KEY") or secrets.token_urlsafe(32),
        SQLALCHEMY_DATABASE_URI=configured_database_uri(app),
        SQLALCHEMY_TRACK_MODIFICATIONS=False,
        JSON_SORT_KEYS=False,
        SESSION_DURATION_MINUTES=int(os.getenv("SESSION_DURATION_MINUTES", "30")),
        MAX_CONTENT_LENGTH=int(os.getenv("MAX_UPLOAD_BYTES", str(DEFAULT_UPLOAD_LIMIT_BYTES))),
    )
    if test_config:
        app.config.update(test_config)

    db.init_app(app)

    origins = [item.strip() for item in os.getenv("CORS_ORIGINS", DEFAULT_CORS_ORIGINS).split(",") if item.strip()]
    CORS(app, resources={r"/api/*": {"origins": origins or "*"}})

    register_error_handlers(app)
    register_security_headers(app)
    register_routes(app)

    with app.app_context():
        db.create_all()
        seed_default_users()

    return app


def seed_default_users() -> None:
    created_any = False
    allow_weak_passwords = bool(current_app.config.get("TESTING")) or os.getenv(
        "ALLOW_INSECURE_INITIAL_PASSWORDS", ""
    ).lower() in {"1", "true", "yes"}

    for username, env_name, role in INITIAL_USERS:
        if User.query.get(username):
            continue
        password = os.getenv(env_name, "").strip()
        if not password:
            current_app.logger.warning("Initial user %s was not created because %s is not set.", username, env_name)
            continue
        if password in WEAK_INITIAL_PASSWORDS and not allow_weak_passwords:
            current_app.logger.warning(
                "Initial user %s was not created because %s uses an insecure default password.",
                username,
                env_name,
            )
            continue
        db.session.add(
            User(
                username=username,
                password_hash=generate_password_hash(password),
                role=role,
                active=True,
                created_at=utcnow(),
            )
        )
        created_any = True
    if created_any:
        db.session.commit()


def hash_token(token: str) -> str:
    return hashlib.sha256(token.encode("utf-8")).hexdigest()


def issue_token(username: str, duration_minutes: int) -> str:
    token = secrets.token_urlsafe(32)
    db.session.add(
        Session(
            token_hash=hash_token(token),
            username=username,
            expires_at=utcnow() + timedelta(minutes=duration_minutes),
            created_at=utcnow(),
            last_seen_at=utcnow(),
        )
    )
    db.session.commit()
    return token


def get_token_from_request(data: dict | None = None) -> str:
    auth_header = request.headers.get("Authorization", "")
    if auth_header.lower().startswith("bearer "):
        return auth_header[7:].strip()
    if data and data.get("token"):
        return str(data["token"]).strip()
    return ""


def find_session_user(token: str, extend: bool = True) -> User | None:
    if not token:
        return None

    session = Session.query.filter_by(token_hash=hash_token(token)).first()
    if not session or not session.user or not session.user.active:
        return None

    now = utcnow()
    if as_utc(session.expires_at) <= now:
        db.session.delete(session)
        db.session.commit()
        return None

    if extend:
        session.expires_at = now + timedelta(minutes=current_session_duration())
        session.last_seen_at = now
        db.session.commit()

    return session.user


def current_session_duration() -> int:
    return int(os.getenv("SESSION_DURATION_MINUTES", "30"))


def api_success(message: str, **extra):
    payload = {"success": True, "message": message}
    payload.update(extra)
    return jsonify(payload)


def api_error(message: str, status: int = 400):
    return jsonify({"success": False, "message": message}), status


def safe_export_cell(value):
    if isinstance(value, str) and (
        value.startswith(FORMULA_LEADING_CHARS) or value.lstrip().startswith(FORMULA_PREFIXES)
    ):
        return "'" + value
    return value


def safe_export_row(values: list) -> list:
    return [safe_export_cell(value) for value in values]


def request_int_arg(name: str, default: int, minimum: int, maximum: int) -> int:
    raw = request.args.get(name, str(default)).strip()
    try:
        value = int(raw)
    except ValueError as exc:
        raise ValueError(f"El parametro {name} debe ser numerico.") from exc
    return min(max(value, minimum), maximum)


def upload_limit_mb() -> str:
    limit = int(current_app.config.get("MAX_CONTENT_LENGTH") or DEFAULT_UPLOAD_LIMIT_BYTES)
    return f"{limit / (1024 * 1024):g}"


def validate_password_strength(password: str) -> None:
    if len(password) < 8:
        raise ValueError("La contrasena debe tener al menos 8 caracteres.")
    if password in WEAK_INITIAL_PASSWORDS:
        raise ValueError("La contrasena es demasiado comun. Use una contrasena unica.")
    if not re.search(r"[A-Za-z]", password) or not re.search(r"\d", password):
        raise ValueError("La contrasena debe incluir letras y numeros.")


def client_ip() -> str:
    forwarded_for = request.headers.get("X-Forwarded-For", "")
    if forwarded_for:
        return forwarded_for.split(",", 1)[0].strip()
    return request.remote_addr or ""


def audit_log_to_dict(log: AuditLog) -> dict:
    details = {}
    if log.details:
        try:
            details = json.loads(log.details)
        except json.JSONDecodeError:
            details = {"raw": log.details}
    return {
        "id": log.id,
        "timestamp": to_timestamp_ms(log.timestamp),
        "timestampIso": as_utc(log.timestamp).isoformat(),
        "actor": log.actor or "",
        "action": log.action,
        "targetType": log.target_type or "",
        "targetId": log.target_id or "",
        "ipAddress": log.ip_address or "",
        "userAgent": log.user_agent or "",
        "details": details,
    }


def add_audit_log(
    action: str,
    target_type: str = "",
    target_id: str | int = "",
    details: dict | None = None,
    actor: str | None = None,
    commit: bool = False,
) -> None:
    current_user = getattr(request, "current_user", None)
    payload = json.dumps(details or {}, ensure_ascii=False)
    db.session.add(
        AuditLog(
            actor=actor or (current_user.username if current_user else None),
            action=action,
            target_type=target_type,
            target_id=str(target_id) if target_id not in (None, "") else "",
            ip_address=client_ip(),
            user_agent=(request.headers.get("User-Agent", "")[:255] or ""),
            details=payload,
        )
    )
    if commit:
        db.session.commit()


def login_rate_limit_config() -> tuple[int, int]:
    max_attempts = int(os.getenv("LOGIN_RATE_LIMIT_MAX_ATTEMPTS", "8"))
    window_minutes = int(os.getenv("LOGIN_RATE_LIMIT_WINDOW_MINUTES", "15"))
    return max_attempts, window_minutes


def login_rate_limit_key(username: str) -> str:
    return f"{client_ip()}:{username.lower()}"


def login_is_limited(username: str) -> bool:
    max_attempts, window_minutes = login_rate_limit_config()
    key = login_rate_limit_key(username)
    cutoff = utcnow() - timedelta(minutes=window_minutes)
    attempts = [item for item in LOGIN_ATTEMPTS.get(key, []) if as_utc(item) >= cutoff]
    LOGIN_ATTEMPTS[key] = attempts
    return len(attempts) >= max_attempts


def register_failed_login(username: str) -> None:
    key = login_rate_limit_key(username)
    LOGIN_ATTEMPTS.setdefault(key, []).append(utcnow())


def clear_failed_logins(username: str) -> None:
    LOGIN_ATTEMPTS.pop(login_rate_limit_key(username), None)


def register_security_headers(app: Flask) -> None:
    @app.after_request
    def add_security_headers(response):
        response.headers.setdefault("X-Content-Type-Options", "nosniff")
        response.headers.setdefault("X-Frame-Options", "DENY")
        response.headers.setdefault("Referrer-Policy", "strict-origin-when-cross-origin")
        response.headers.setdefault("Permissions-Policy", "camera=(), microphone=(), geolocation=()")
        response.headers.setdefault(
            "Content-Security-Policy",
            "default-src 'self'; "
            "script-src 'self' 'unsafe-inline'; "
            "style-src 'self' 'unsafe-inline'; "
            "img-src 'self' data:; "
            "connect-src 'self' https://cisternas-api-wqac.onrender.com; "
            "base-uri 'self'; "
            "form-action 'self'; "
            "frame-ancestors 'none'",
        )
        if request.is_secure or request.headers.get("X-Forwarded-Proto", "").lower() == "https":
            response.headers.setdefault("Strict-Transport-Security", "max-age=31536000; includeSubDomains")
        return response


def register_error_handlers(app: Flask) -> None:
    @app.errorhandler(RequestEntityTooLarge)
    def handle_request_too_large(_error):
        return api_error(f"Archivo demasiado grande. Maximo permitido: {upload_limit_mb()} MB.", 413)


def require_auth(admin: bool = False):
    def decorator(func):
        @wraps(func)
        def wrapper(*args, **kwargs):
            user = find_session_user(get_token_from_request(), extend=True)
            if user is None:
                return api_error("Sesion invalida o expirada.", 401)
            if admin and user.role != "admin":
                return api_error("No tiene permisos de administrador.", 403)
            request.current_user = user
            return func(*args, **kwargs)

        return wrapper

    return decorator


def parse_float(data: dict, key: str) -> float | None:
    try:
        value = data.get(key)
        if value is None or value == "":
            return None
        return float(value)
    except (TypeError, ValueError):
        return None


def clean_text(data: dict, key: str) -> str:
    return str(data.get(key, "")).strip()


def uppercase_text(value) -> str:
    text = str(value or "").strip().upper()
    return re.sub(r"\s+", " ", text)


def normalize_plate_number(value) -> str:
    raw = re.sub(r"[^A-Za-z0-9]", "", str(value or "")).upper()
    match = re.fullmatch(r"(\d{4})([A-Z]{3})", raw)
    if not match:
        raise ValueError("La placa debe tener el formato 2127 - ACC.")
    return f"{match.group(1)} - {match.group(2)}"


def parse_filter_date(value: str, add_day: bool = False) -> datetime | None:
    if not value:
        return None
    try:
        parsed = datetime.fromisoformat(value)
    except ValueError:
        return None
    if parsed.tzinfo is None:
        parsed = parsed.replace(tzinfo=timezone.utc)
    parsed = parsed.astimezone(timezone.utc)
    if add_day:
        parsed = parsed + timedelta(days=1)
    return parsed


def record_to_dict(record: WaterRecord) -> dict:
    timestamp = as_utc(record.timestamp)
    return {
        "id": record.id,
        "timestamp": to_timestamp_ms(timestamp),
        "timestampIso": timestamp.isoformat(),
        "driverName": record.driver_name,
        "plateNumber": record.plate_number,
        "employeeCode": record.employee_code,
        "ebap": record.ebap,
        "initialReading": record.initial_reading,
        "finalReading": record.final_reading,
        "loadVolume": record.load_volume,
        "companyType": record.company_type,
        "companyName": record.company_name,
        "characteristics": record.characteristics or "",
        "registeredBy": record.registered_by,
    }


def user_to_dict(user: User) -> dict:
    return {
        "username": user.username,
        "role": user.role,
        "active": user.active,
        "createdAt": as_utc(user.created_at).isoformat(),
    }


def record_export_row(record: WaterRecord) -> list:
    return safe_export_row([
        record.id,
        as_utc(record.timestamp).isoformat(),
        record.driver_name,
        record.plate_number,
        record.employee_code,
        record.ebap,
        record.initial_reading,
        record.final_reading,
        record.load_volume,
        record.company_type,
        record.company_name,
        record.characteristics or "",
        record.registered_by,
    ])


def audit_export_row(log: AuditLog) -> list:
    return safe_export_row([
        log.id,
        as_utc(log.timestamp).isoformat(),
        log.actor or "",
        log.action,
        log.target_type or "",
        log.target_id or "",
        log.ip_address or "",
        log.user_agent or "",
        log.details or "",
    ])


def set_sheet_widths(sheet, widths: list[int]) -> None:
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[chr(64 + index)].width = width


def normalize_header(value) -> str:
    text = unicodedata.normalize("NFKD", str(value or ""))
    text = "".join(char for char in text if not unicodedata.combining(char))
    return "".join(char.lower() for char in text if char.isalnum())


def normalized_import_row(row: dict) -> dict:
    return {normalize_header(key): value for key, value in row.items() if key is not None}


def import_value(row: dict, field: str, default: str = ""):
    aliases = IMPORT_ALIASES[field]
    for alias in aliases:
        if alias in row and row[alias] not in (None, ""):
            return row[alias]
    return default


def parse_import_float(value, field_label: str) -> float:
    if value is None or value == "":
        raise ValueError(f"{field_label} requerido.")
    try:
        return float(str(value).replace(",", ".").strip())
    except (TypeError, ValueError) as exc:
        raise ValueError(f"{field_label} no es numerico.") from exc


def parse_import_timestamp(value) -> datetime:
    if isinstance(value, datetime):
        return as_utc(value)
    if value is None or value == "":
        return utcnow()

    raw = str(value).strip()
    if not raw:
        return utcnow()

    try:
        numeric = float(raw)
        if numeric > 1_000_000_000_000:
            return datetime.fromtimestamp(numeric / 1000, tz=timezone.utc)
        if numeric > 1_000_000_000:
            return datetime.fromtimestamp(numeric, tz=timezone.utc)
    except ValueError:
        pass

    iso_value = raw.replace("Z", "+00:00")
    try:
        return as_utc(datetime.fromisoformat(iso_value))
    except ValueError:
        pass

    for fmt in ("%d/%m/%Y %H:%M", "%d/%m/%y %H:%M", "%d/%m/%Y", "%d/%m/%y"):
        try:
            return datetime.strptime(raw, fmt).replace(tzinfo=timezone.utc)
        except ValueError:
            continue

    return utcnow()


def read_csv_import_rows(file_storage) -> list[dict]:
    raw = file_storage.read()
    try:
        text = raw.decode("utf-8-sig")
    except UnicodeDecodeError:
        text = raw.decode("latin-1")
    return list(csv.DictReader(io.StringIO(text)))


def read_xlsx_import_rows(file_storage) -> list[dict]:
    file_storage.stream.seek(0)
    workbook = load_workbook(io.BytesIO(file_storage.read()), read_only=True, data_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(value or "").strip() for value in rows[0]]
    result = []
    for row_values in rows[1:]:
        if not any(value not in (None, "") for value in row_values):
            continue
        result.append({headers[index]: row_values[index] if index < len(row_values) else "" for index in range(len(headers))})
    return result


def read_import_rows(file_storage) -> list[dict]:
    filename = (file_storage.filename or "").lower()
    if filename.endswith(".xlsx"):
        return read_xlsx_import_rows(file_storage)
    if filename.endswith(".csv"):
        return read_csv_import_rows(file_storage)
    raise ValueError("Formato no soportado. Use CSV o XLSX.")


def build_import_record(row: dict, fallback_username: str) -> WaterRecord:
    normalized = normalized_import_row(row)

    driver_name = uppercase_text(import_value(normalized, "driver_name"))
    plate_number = normalize_plate_number(import_value(normalized, "plate_number"))
    employee_code = uppercase_text(import_value(normalized, "employee_code"))
    ebap = uppercase_text(import_value(normalized, "ebap"))
    company_type = uppercase_text(import_value(normalized, "company_type"))
    company_name = uppercase_text(import_value(normalized, "company_name"))
    characteristics = uppercase_text(import_value(normalized, "characteristics"))

    if not company_type and " / " in company_name:
        company_type, company_name = [part.strip() for part in company_name.split(" / ", 1)]

    required = {
        "Conductor": driver_name,
        "Placa": plate_number,
        "Codigo funcionario": employee_code,
        "EBAP": ebap,
        "Tipo empresa": company_type,
        "Empresa/Entidad": company_name,
    }
    missing = [label for label, value in required.items() if not value]
    if missing:
        raise ValueError("Faltan campos: " + ", ".join(missing) + ".")

    initial = parse_import_float(import_value(normalized, "initial_reading"), "Lectura inicial")
    final = parse_import_float(import_value(normalized, "final_reading"), "Lectura final")
    if initial < 0 or final < 0 or final <= initial:
        raise ValueError("Lecturas invalidas.")

    volume_value = import_value(normalized, "load_volume", "")
    load_volume = parse_import_float(volume_value, "Volumen") if volume_value not in (None, "") else round(final - initial, 2)
    if load_volume <= 0:
        raise ValueError("Volumen invalido.")

    registered_by = str(import_value(normalized, "registered_by", fallback_username)).strip() or fallback_username
    if not User.query.get(registered_by):
        registered_by = fallback_username

    return WaterRecord(
        timestamp=parse_import_timestamp(import_value(normalized, "timestamp", "")),
        driver_name=driver_name,
        plate_number=plate_number,
        employee_code=employee_code,
        ebap=ebap,
        initial_reading=initial,
        final_reading=final,
        load_volume=round(load_volume, 2),
        company_type=company_type,
        company_name=company_name,
        characteristics=characteristics,
        registered_by=registered_by,
    )


def validated_record_values(data: dict) -> dict:
    try:
        plate_number = normalize_plate_number(clean_text(data, "plateNumber"))
    except ValueError as exc:
        raise ValueError(str(exc)) from exc

    required_text = {
        "driverName": uppercase_text(data.get("driverName")),
        "plateNumber": plate_number,
        "employeeCode": uppercase_text(data.get("employeeCode")),
        "ebap": uppercase_text(data.get("ebap")),
        "companyType": uppercase_text(data.get("companyType")),
        "companyName": uppercase_text(data.get("companyName")),
    }
    if any(not value for value in required_text.values()):
        raise ValueError("Complete todos los campos obligatorios.")

    initial = parse_float(data, "initialReading")
    final = parse_float(data, "finalReading")
    volume = parse_float(data, "loadVolume")
    if initial is None or final is None or initial < 0 or final < 0 or final <= initial:
        raise ValueError("Las lecturas no son validas.")

    calculated_volume = round(final - initial, 2)
    if volume is None or volume <= 0:
        volume = calculated_volume
    if volume <= 0:
        raise ValueError("El volumen calculado debe ser mayor a cero.")

    return {
        "driver_name": required_text["driverName"],
        "plate_number": required_text["plateNumber"],
        "employee_code": required_text["employeeCode"],
        "ebap": required_text["ebap"],
        "initial_reading": initial,
        "final_reading": final,
        "load_volume": round(volume, 2),
        "company_type": required_text["companyType"],
        "company_name": required_text["companyName"],
        "characteristics": uppercase_text(data.get("characteristics")),
    }


def login_action(data: dict):
    username = clean_text(data, "username")
    password = str(data.get("password", ""))
    if not username or not password:
        return api_error("Usuario y contrasena requeridos.")

    if login_is_limited(username):
        add_audit_log(
            "login_rate_limited",
            "user",
            username,
            {"reason": "too_many_failed_attempts"},
            actor=username,
            commit=True,
        )
        return api_error("Demasiados intentos fallidos. Intente nuevamente mas tarde.", 429)

    user = User.query.get(username)
    if not user or not user.active or not check_password_hash(user.password_hash, password):
        register_failed_login(username)
        add_audit_log(
            "login_failed",
            "user",
            username,
            {"reason": "invalid_credentials"},
            actor=username,
            commit=True,
        )
        return api_error("Credenciales incorrectas.", 401)

    clear_failed_logins(username)
    token = issue_token(user.username, current_session_duration())
    add_audit_log("login_success", "user", user.username, actor=user.username, commit=True)
    return api_success(
        "Login exitoso",
        token=token,
        user={"username": user.username, "role": user.role},
    )


def validate_token_action(data: dict):
    user = find_session_user(get_token_from_request(data), extend=True)
    if user is None:
        return api_error("Token invalido o expirado.", 401)
    return api_success("Token valido", user={"username": user.username, "role": user.role})


def save_record_action(data: dict):
    user = find_session_user(get_token_from_request(data), extend=True)
    if user is None:
        return api_error("Sesion expirada. Inicie sesion nuevamente.", 401)

    try:
        values = validated_record_values(data)
    except ValueError as exc:
        return api_error(str(exc))

    record = WaterRecord(
        timestamp=utcnow(),
        **values,
        registered_by=user.username,
    )
    db.session.add(record)
    db.session.flush()
    add_audit_log(
        "record_created",
        "record",
        record.id,
        {"registeredBy": user.username, "plateNumber": record.plate_number},
        actor=user.username,
    )
    db.session.commit()
    return api_success("Datos guardados en la base central", id=record.id)


def list_records_for_user(data: dict):
    user = find_session_user(get_token_from_request(data), extend=True)
    if user is None:
        return api_error("Sesion expirada. Inicie sesion nuevamente.", 401)

    query = WaterRecord.query
    if user.role != "admin":
        query = query.filter(WaterRecord.registered_by == user.username)
    records = query.order_by(WaterRecord.id.desc()).limit(500).all()
    return api_success("Registros cargados", records=[record_to_dict(item) for item in records])


def apply_record_filters(query):
    q = request.args.get("q", "").strip()
    ebap = request.args.get("ebap", "").strip()
    company_type = request.args.get("companyType", "").strip()
    date_from = request.args.get("dateFrom", "").strip()
    date_to = request.args.get("dateTo", "").strip()

    if q:
        pattern = f"%{q}%"
        query = query.filter(
            or_(
                WaterRecord.driver_name.ilike(pattern),
                WaterRecord.plate_number.ilike(pattern),
                WaterRecord.employee_code.ilike(pattern),
                WaterRecord.company_name.ilike(pattern),
                WaterRecord.registered_by.ilike(pattern),
            )
        )
    if ebap:
        query = query.filter(WaterRecord.ebap.ilike(ebap))
    if company_type:
        query = query.filter(WaterRecord.company_type.ilike(company_type))
    start = parse_filter_date(date_from)
    if start:
        query = query.filter(WaterRecord.timestamp >= start)
    end = parse_filter_date(date_to, add_day=True)
    if end:
        query = query.filter(WaterRecord.timestamp < end)
    return query


def register_routes(app: Flask) -> None:
    @app.get("/")
    def index():
        return redirect("/admin")

    @app.get("/admin")
    def admin_panel():
        return render_template("admin.html")

    @app.get("/app")
    @app.get("/app/")
    def mobile_web_app():
        return send_from_directory(MOBILE_APP_DIR, "index.html")

    @app.get("/app/<path:filename>")
    def mobile_web_asset(filename: str):
        return send_from_directory(MOBILE_APP_DIR, filename)

    @app.get("/api/health")
    def health():
        return api_success("API operativa")

    @app.post("/api/mobile")
    def mobile_api():
        data = request.get_json(silent=True) or {}
        action = str(data.get("action", "")).strip()
        if action == "login":
            return login_action(data)
        if action == "validateToken":
            return validate_token_action(data)
        if action == "saveData":
            return save_record_action(data)
        if action == "listData":
            return list_records_for_user(data)
        if action == "test":
            return api_success("API central operativa")
        return api_error("Accion no valida.")

    @app.post("/api/auth/login")
    def api_login():
        return login_action(request.get_json(silent=True) or {})

    @app.post("/api/auth/validate")
    def api_validate():
        return validate_token_action(request.get_json(silent=True) or {})

    @app.post("/api/auth/logout")
    def api_logout():
        token = get_token_from_request(request.get_json(silent=True) or {})
        if token:
            Session.query.filter_by(token_hash=hash_token(token)).delete()
            db.session.commit()
        return api_success("Sesion cerrada")

    @app.get("/api/records")
    @require_auth(admin=True)
    def api_records():
        try:
            limit = request_int_arg("limit", 500, 1, 2000)
        except ValueError as exc:
            return api_error(str(exc))
        query = apply_record_filters(WaterRecord.query)
        records = query.order_by(WaterRecord.id.desc()).limit(limit).all()
        total_volume = round(sum(item.load_volume for item in records), 2)
        return api_success(
            "Registros cargados",
            records=[record_to_dict(item) for item in records],
            total=len(records),
            totalVolume=total_volume,
        )

    @app.patch("/api/records/<int:record_id>")
    @require_auth(admin=True)
    def api_update_record(record_id: int):
        record = WaterRecord.query.get(record_id)
        if not record:
            return api_error("Registro no encontrado.", 404)

        before = record_to_dict(record)
        try:
            values = validated_record_values(request.get_json(silent=True) or {})
        except ValueError as exc:
            return api_error(str(exc))

        for field, value in values.items():
            setattr(record, field, value)

        add_audit_log(
            "record_updated",
            "record",
            record.id,
            {"before": before, "after": record_to_dict(record)},
        )
        db.session.commit()
        return api_success("Registro actualizado", record=record_to_dict(record))

    @app.delete("/api/records/<int:record_id>")
    @require_auth(admin=True)
    def api_delete_record(record_id: int):
        record = WaterRecord.query.get(record_id)
        if not record:
            return api_error("Registro no encontrado.", 404)

        snapshot = record_to_dict(record)
        db.session.delete(record)
        add_audit_log("record_deleted", "record", record_id, {"record": snapshot})
        db.session.commit()
        return api_success("Registro eliminado")

    @app.get("/api/records/export.csv")
    @require_auth(admin=True)
    def export_records_csv():
        query = apply_record_filters(WaterRecord.query)
        records = query.order_by(WaterRecord.id.desc()).limit(10000).all()

        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow(RECORD_EXPORT_HEADERS)
        for record in records:
            writer.writerow(record_export_row(record))

        return Response(
            output.getvalue(),
            mimetype="text/csv; charset=utf-8",
            headers={"Content-Disposition": "attachment; filename=registros_cisternas.csv"},
        )

    @app.get("/api/records/export.xlsx")
    @require_auth(admin=True)
    def export_records_xlsx():
        query = apply_record_filters(WaterRecord.query)
        records = query.order_by(WaterRecord.id.desc()).limit(10000).all()

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Registros"
        sheet.append(RECORD_EXPORT_HEADERS)
        for record in records:
            sheet.append(record_export_row(record))

        set_sheet_widths(sheet, [8, 24, 26, 14, 18, 14, 16, 16, 12, 24, 28, 34, 18])

        output = io.BytesIO()
        workbook.save(output)
        output.seek(0)
        return Response(
            output.getvalue(),
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=registros_cisternas.xlsx"},
        )

    @app.get("/api/backup.xlsx")
    @require_auth(admin=True)
    def export_full_backup_xlsx():
        generated_at = utcnow()
        filename = f"respaldo_cisternas_scpe_{generated_at.strftime('%Y%m%d_%H%M%S')}.xlsx"

        records = WaterRecord.query.order_by(WaterRecord.id.asc()).all()
        users = User.query.order_by(User.username.asc()).all()

        add_audit_log(
            "backup_exported",
            "backup",
            filename,
            {"records": len(records), "users": len(users)},
        )
        db.session.commit()

        audit_logs = AuditLog.query.order_by(AuditLog.id.asc()).all()

        workbook = Workbook()
        summary = workbook.active
        summary.title = "Resumen"
        summary.append(["Campo", "Valor"])
        summary.append(["Generado UTC", as_utc(generated_at).isoformat()])
        summary.append(["Generado por", request.current_user.username])
        summary.append(["Registros incluidos", len(records)])
        summary.append(["Usuarios incluidos", len(users)])
        summary.append(["Eventos de auditoria incluidos", len(audit_logs)])
        summary.append(["Nota", "Respaldo manual descargado desde el panel administrador SCPE."])
        set_sheet_widths(summary, [32, 80])

        records_sheet = workbook.create_sheet("Registros")
        records_sheet.append(RECORD_EXPORT_HEADERS)
        for record in records:
            records_sheet.append(record_export_row(record))
        set_sheet_widths(records_sheet, [8, 24, 26, 14, 18, 14, 16, 16, 12, 24, 28, 34, 18])

        users_sheet = workbook.create_sheet("Usuarios")
        users_sheet.append(["Usuario", "Rol", "Activo", "Creado UTC"])
        for user in users:
            users_sheet.append([user.username, user.role, "SI" if user.active else "NO", as_utc(user.created_at).isoformat()])
        set_sheet_widths(users_sheet, [24, 18, 12, 24])

        audit_sheet = workbook.create_sheet("Auditoria")
        audit_sheet.append(["ID", "Fecha UTC", "Usuario", "Accion", "Tipo objetivo", "ID objetivo", "IP", "User-Agent", "Detalle"])
        for log in audit_logs:
            audit_sheet.append(audit_export_row(log))
        set_sheet_widths(audit_sheet, [8, 24, 20, 24, 18, 24, 18, 36, 80])

        output = io.BytesIO()
        workbook.save(output)
        output.seek(0)
        return Response(
            output.getvalue(),
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"},
        )

    @app.get("/api/records/export.pdf")
    @require_auth(admin=True)
    def export_records_pdf():
        query = apply_record_filters(WaterRecord.query)
        records = query.order_by(WaterRecord.id.desc()).limit(2000).all()

        output = io.BytesIO()
        doc = SimpleDocTemplate(
            output,
            pagesize=landscape(letter),
            rightMargin=24,
            leftMargin=24,
            topMargin=24,
            bottomMargin=24,
        )
        styles = getSampleStyleSheet()
        elements = [
            Paragraph("Registro de carguio de cisternas - SCPE", styles["Title"]),
            Paragraph(f"Registros exportados: {len(records)}", styles["Normal"]),
            Spacer(1, 12),
        ]

        pdf_headers = ["ID", "Fecha", "Conductor", "Placa", "Func.", "EBAP", "Lecturas", "Vol.", "Empresa", "Obs.", "Usuario"]
        table_data = [pdf_headers]
        for record in records:
            row = [
                record.id,
                as_utc(record.timestamp).strftime("%d/%m/%Y %H:%M"),
                record.driver_name,
                record.plate_number,
                record.employee_code,
                record.ebap,
                f"{record.initial_reading:g} -> {record.final_reading:g}",
                f"{record.load_volume:.2f}",
                f"{record.company_type} / {record.company_name}",
                record.characteristics or "",
                record.registered_by,
            ]
            table_data.append([Paragraph(str(value), styles["BodyText"]) for value in row])

        table = Table(table_data, repeatRows=1, colWidths=[28, 76, 90, 58, 52, 58, 68, 44, 130, 120, 58])
        table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0b3f59")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                    ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#cfe0ed")),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("FONTSIZE", (0, 0), (-1, -1), 7),
                    ("VALIGN", (0, 0), (-1, -1), "TOP"),
                    ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f1fbfd")]),
                ]
            )
        )
        elements.append(table)
        doc.build(elements)
        output.seek(0)
        return Response(
            output.getvalue(),
            mimetype="application/pdf",
            headers={"Content-Disposition": "attachment; filename=registros_cisternas.pdf"},
        )

    @app.post("/api/records/import")
    @require_auth(admin=True)
    def import_records():
        file_storage = request.files.get("file")
        if file_storage is None or not file_storage.filename:
            return api_error("Seleccione un archivo CSV o XLSX.")

        try:
            rows = read_import_rows(file_storage)
        except ValueError as exc:
            return api_error(str(exc))

        imported = 0
        skipped = 0
        errors = []
        for index, row in enumerate(rows, start=2):
            try:
                db.session.add(build_import_record(row, request.current_user.username))
                imported += 1
            except ValueError as exc:
                skipped += 1
                if len(errors) < 20:
                    errors.append(f"Fila {index}: {exc}")

        if imported:
            db.session.commit()
            add_audit_log(
                "records_imported",
                "record",
                "bulk",
                {"imported": imported, "skipped": skipped, "errors": errors},
                commit=True,
            )
        else:
            db.session.rollback()

        return api_success(
            "Importacion finalizada",
            imported=imported,
            skipped=skipped,
            errors=errors,
        )

    @app.get("/api/users")
    @require_auth(admin=True)
    def api_users():
        users = User.query.order_by(User.username.asc()).all()
        return api_success("Usuarios cargados", users=[user_to_dict(user) for user in users])

    @app.post("/api/users")
    @require_auth(admin=True)
    def api_create_user():
        data = request.get_json(silent=True) or {}
        username = clean_text(data, "username").lower()
        password = str(data.get("password", ""))
        role = clean_text(data, "role") or "user"

        if not username or not password:
            return api_error("Usuario y contrasena requeridos.")
        if len(username) < 3:
            return api_error("El usuario debe tener al menos 3 caracteres.")
        try:
            validate_password_strength(password)
        except ValueError as exc:
            return api_error(str(exc))
        if role not in {"admin", "user"}:
            return api_error("Rol no valido.")
        if User.query.get(username):
            return api_error("Ese usuario ya existe.")

        user = User(
            username=username,
            password_hash=generate_password_hash(password),
            role=role,
            active=True,
            created_at=utcnow(),
        )
        db.session.add(user)
        add_audit_log("user_created", "user", user.username, {"role": user.role, "active": user.active})
        db.session.commit()
        return api_success("Usuario creado", user=user_to_dict(user))

    @app.patch("/api/users/<username>/password")
    @require_auth(admin=True)
    def api_update_user_password(username: str):
        data = request.get_json(silent=True) or {}
        target_username = username.strip().lower()
        user = User.query.get(target_username)
        if not user:
            return api_error("Usuario no encontrado.", 404)

        password = str(data.get("password", ""))
        if not password:
            return api_error("Nueva contrasena requerida.")
        try:
            validate_password_strength(password)
        except ValueError as exc:
            return api_error(str(exc))

        current_user = request.current_user
        current_token = get_token_from_request()
        current_token_hash = hash_token(current_token) if current_token else ""
        if user.username == current_user.username:
            current_password = str(data.get("currentPassword", ""))
            if not current_password or not check_password_hash(user.password_hash, current_password):
                return api_error("Contrasena actual incorrecta.", 403)

        user.password_hash = generate_password_hash(password)
        sessions_query = Session.query.filter(Session.username == user.username)
        if user.username == current_user.username and current_token_hash:
            sessions_query = sessions_query.filter(Session.token_hash != current_token_hash)
        revoked_sessions = sessions_query.delete(synchronize_session=False)
        add_audit_log(
            "user_password_changed",
            "user",
            user.username,
            {"self": user.username == current_user.username, "revokedSessions": revoked_sessions},
        )
        db.session.commit()
        return api_success("Contrasena actualizada", revokedSessions=revoked_sessions)

    @app.patch("/api/users/<username>")
    @require_auth(admin=True)
    def api_update_user(username: str):
        data = request.get_json(silent=True) or {}
        target_username = username.strip().lower()
        user = User.query.get(target_username)
        if not user:
            return api_error("Usuario no encontrado.", 404)

        current_user = request.current_user
        if user.username == current_user.username and data.get("active") is False:
            return api_error("No puede desactivar su propio usuario administrador.")

        before = user_to_dict(user)
        if "active" in data:
            user.active = bool(data["active"])

        if "role" in data:
            role = clean_text(data, "role")
            if role not in {"admin", "user"}:
                return api_error("Rol no valido.")
            if user.username == current_user.username and role != "admin":
                return api_error("No puede quitarse su propio rol administrador.")
            user.role = role

        add_audit_log("user_updated", "user", user.username, {"before": before, "after": user_to_dict(user)})
        db.session.commit()
        return api_success("Usuario actualizado", user=user_to_dict(user))

    @app.get("/api/audit-logs")
    @require_auth(admin=True)
    def api_audit_logs():
        try:
            limit = request_int_arg("limit", 200, 1, 1000)
        except ValueError as exc:
            return api_error(str(exc))
        action = request.args.get("action", "").strip()
        actor = request.args.get("actor", "").strip()

        query = AuditLog.query
        if action:
            query = query.filter(AuditLog.action == action)
        if actor:
            query = query.filter(AuditLog.actor.ilike(f"%{actor}%"))

        logs = query.order_by(AuditLog.id.desc()).limit(limit).all()
        return api_success("Auditoria cargada", logs=[audit_log_to_dict(item) for item in logs])


app = create_app()


if __name__ == "__main__":
    debug_enabled = os.getenv("FLASK_DEBUG", "false").lower() in {"1", "true", "yes"}
    app.run(host=os.getenv("HOST", "127.0.0.1"), port=int(os.getenv("PORT", "5000")), debug=debug_enabled)
