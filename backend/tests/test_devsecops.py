from __future__ import annotations

import csv
from io import BytesIO
from io import StringIO

import pytest
from openpyxl import load_workbook

from app import LOGIN_ATTEMPTS, User, create_app

ADMIN_TEST_PASSWORD = "TestAdminPassword123!"
USER_TEST_PASSWORD = "TestUserPassword123!"
CLIENT_TEST_PASSWORD = "TestClientPassword123!"
NEW_CLIENT_TEST_PASSWORD = "NewClientPassword987!"
NEW_ADMIN_TEST_PASSWORD = "NewAdminPassword987!"


@pytest.fixture()
def client(tmp_path, monkeypatch):
    LOGIN_ATTEMPTS.clear()
    monkeypatch.setenv("ADMIN_PASSWORD", ADMIN_TEST_PASSWORD)
    monkeypatch.setenv("USER_PASSWORD", USER_TEST_PASSWORD)
    monkeypatch.setenv("CLIENT_PASSWORD", CLIENT_TEST_PASSWORD)

    app = create_app(
        {
            "TESTING": True,
            "SQLALCHEMY_DATABASE_URI": "sqlite:///" + str(tmp_path / "test.db"),
            "SESSION_DURATION_MINUTES": 30,
        }
    )
    return app.test_client()


def ok(response):
    data = response.get_json()
    assert response.status_code == 200, (response.status_code, data)
    assert data["success"] is True, data
    return data


def login(client, username: str, password: str) -> dict:
    return ok(client.post("/api/auth/login", json={"username": username, "password": password}))


def auth_header(token: str) -> dict:
    return {"Authorization": f"Bearer {token}"}


def test_security_headers_are_present(client):
    response = client.get("/api/health", headers={"X-Forwarded-Proto": "https"})

    assert response.headers["X-Content-Type-Options"] == "nosniff"
    assert response.headers["X-Frame-Options"] == "DENY"
    assert "frame-ancestors 'none'" in response.headers["Content-Security-Policy"]
    assert "Strict-Transport-Security" in response.headers


def test_cors_is_restricted_to_expected_origins(client):
    allowed = client.get("/api/health", headers={"Origin": "https://cisternas-api-wqac.onrender.com"})
    null_origin = client.get("/api/health", headers={"Origin": "null"})
    denied = client.get("/api/health", headers={"Origin": "https://evil.example"})

    assert allowed.headers["Access-Control-Allow-Origin"] == "https://cisternas-api-wqac.onrender.com"
    assert null_origin.headers["Access-Control-Allow-Origin"] == "null"
    assert "Access-Control-Allow-Origin" not in denied.headers


def test_insecure_initial_passwords_are_not_seeded_outside_testing(tmp_path, monkeypatch):
    monkeypatch.setenv("ADMIN_PASSWORD", "admin123")
    monkeypatch.setenv("USER_PASSWORD", "usuario123")
    monkeypatch.setenv("CLIENT_PASSWORD", "cliente123")

    app = create_app({"SQLALCHEMY_DATABASE_URI": "sqlite:///" + str(tmp_path / "prod_like.db")})

    with app.app_context():
        assert User.query.get("admin") is None
        assert User.query.get("usuario") is None
        assert User.query.get("cliente") is None


def test_records_are_normalized_and_invalid_plate_is_rejected(client):
    admin = login(client, "admin", ADMIN_TEST_PASSWORD)
    user = login(client, "cliente", CLIENT_TEST_PASSWORD)

    created = ok(
        client.post(
            "/api/mobile",
            json={
                "action": "saveData",
                "token": user["token"],
                "driverName": "cristian montoya",
                "plateNumber": "2127acc",
                "employeeCode": "emp 1277",
                "ebap": "norte",
                "initialReading": 10,
                "finalReading": 33,
                "loadVolume": 23,
                "companyType": "gobierno nacional",
                "companyName": "andina sa",
                "characteristics": "volvo color blanco",
            },
        )
    )

    records = ok(client.get("/api/records", headers=auth_header(admin["token"])))["records"]
    record = next(item for item in records if item["id"] == created["id"])
    assert record["driverName"] == "CRISTIAN MONTOYA"
    assert record["plateNumber"] == "2127 - ACC"
    assert record["companyName"] == "ANDINA SA"

    invalid = client.post(
        "/api/mobile",
        json={
            "action": "saveData",
            "token": user["token"],
            "driverName": "x",
            "plateNumber": "21acc",
            "employeeCode": "e",
            "ebap": "norte",
            "initialReading": 1,
            "finalReading": 2,
            "loadVolume": 1,
            "companyType": "particular",
            "companyName": "x",
            "characteristics": "",
        },
    )
    assert invalid.status_code == 400
    assert invalid.get_json()["success"] is False


def test_mobile_record_uses_authenticated_user_when_employee_code_is_omitted(client):
    admin = login(client, "admin", ADMIN_TEST_PASSWORD)
    user = login(client, "cliente", CLIENT_TEST_PASSWORD)

    created = ok(
        client.post(
            "/api/mobile",
            json={
                "action": "saveData",
                "token": user["token"],
                "driverName": "operador sin codigo",
                "plateNumber": "4444abc",
                "ebap": "norte",
                "initialReading": 10,
                "finalReading": 20,
                "loadVolume": 10,
                "companyType": "particular",
                "companyName": "andina",
                "characteristics": "",
            },
        )
    )

    records = ok(client.get("/api/records", headers=auth_header(admin["token"])))["records"]
    record = next(item for item in records if item["id"] == created["id"])
    assert record["employeeCode"] == "CLIENTE"


def test_import_does_not_require_employee_code(client):
    admin = login(client, "admin", ADMIN_TEST_PASSWORD)
    headers = auth_header(admin["token"])
    csv_data = (
        "Conductor,Placa,EBAP,Lectura inicial,Lectura final,Tipo empresa,Empresa/Entidad,Caracteristicas,Registrado por\n"
        "operador importado,1234ABC,norte,10,20,particular,empresa prueba,,cliente\n"
    )

    imported = ok(
        client.post(
            "/api/records/import",
            headers=headers,
            data={"file": (BytesIO(csv_data.encode("utf-8")), "registros.csv")},
            content_type="multipart/form-data",
        )
    )

    records = ok(client.get("/api/records", headers=headers))["records"]
    record = next(item for item in records if item["driverName"] == "OPERADOR IMPORTADO")
    assert imported["imported"] == 1
    assert imported["skipped"] == 0
    assert record["employeeCode"] == "CLIENTE"
    assert record["registeredBy"] == "cliente"


def test_login_rate_limit_blocks_repeated_failures(client, monkeypatch):
    monkeypatch.setenv("LOGIN_RATE_LIMIT_MAX_ATTEMPTS", "2")
    monkeypatch.setenv("LOGIN_RATE_LIMIT_WINDOW_MINUTES", "15")

    first = client.post("/api/auth/login", json={"username": "admin", "password": "bad"})
    second = client.post("/api/auth/login", json={"username": "admin", "password": "bad"})
    third = client.post("/api/auth/login", json={"username": "admin", "password": "bad"})

    assert first.status_code == 401
    assert second.status_code == 401
    assert third.status_code == 429
    assert "Demasiados intentos" in third.get_json()["message"]


def test_admin_actions_are_audited(client):
    admin = login(client, "admin", ADMIN_TEST_PASSWORD)
    user = login(client, "cliente", CLIENT_TEST_PASSWORD)
    headers = auth_header(admin["token"])

    created = ok(
        client.post(
            "/api/mobile",
            json={
                "action": "saveData",
                "token": user["token"],
                "driverName": "jose ortega",
                "plateNumber": "6161acr",
                "employeeCode": "1276",
                "ebap": "este",
                "initialReading": 80,
                "finalReading": 90,
                "loadVolume": 10,
                "companyType": "bomberos",
                "companyName": "andina",
                "characteristics": "volvo rojo",
            },
        )
    )

    ok(
        client.patch(
            f"/api/records/{created['id']}",
            headers=headers,
            json={
                "driverName": "jose editado",
                "plateNumber": "6161acr",
                "employeeCode": "1276",
                "ebap": "este",
                "initialReading": 80,
                "finalReading": 91,
                "loadVolume": 11,
                "companyType": "bomberos",
                "companyName": "andina",
                "characteristics": "editado",
            },
        )
    )
    ok(client.delete(f"/api/records/{created['id']}", headers=headers))

    logs = ok(client.get("/api/audit-logs?limit=50", headers=headers))["logs"]
    actions = {item["action"] for item in logs}
    assert "record_created" in actions
    assert "record_updated" in actions
    assert "record_deleted" in actions


def test_admin_can_download_full_backup(client):
    admin = login(client, "admin", ADMIN_TEST_PASSWORD)

    response = client.get("/api/backup.xlsx", headers=auth_header(admin["token"]))

    assert response.status_code == 200
    assert response.mimetype == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    assert "respaldo_cisternas_scpe_" in response.headers["Content-Disposition"]

    workbook = load_workbook(BytesIO(response.data), read_only=True)
    assert {"Resumen", "Registros", "Usuarios", "Auditoria"}.issubset(set(workbook.sheetnames))
    records_header = [cell.value for cell in next(workbook["Registros"].iter_rows(min_row=1, max_row=1))]
    assert "Codigo funcionario" not in records_header


def test_record_pdf_export_does_not_include_employee_code_header(client):
    admin = login(client, "admin", ADMIN_TEST_PASSWORD)

    response = client.get("/api/records/export.pdf", headers=auth_header(admin["token"]))

    assert response.status_code == 200
    assert b"Func." not in response.data
    assert b"Codigo funcionario" not in response.data


def test_admin_can_change_user_password_and_revoke_sessions(client):
    admin = login(client, "admin", ADMIN_TEST_PASSWORD)
    user = login(client, "cliente", CLIENT_TEST_PASSWORD)
    headers = auth_header(admin["token"])

    changed = ok(
        client.patch(
            "/api/users/cliente/password",
            headers=headers,
            json={"password": NEW_CLIENT_TEST_PASSWORD},
        )
    )

    assert changed["revokedSessions"] >= 1
    old_login = client.post("/api/auth/login", json={"username": "cliente", "password": CLIENT_TEST_PASSWORD})
    old_token = client.post("/api/auth/validate", json={"token": user["token"]})
    new_login = client.post("/api/auth/login", json={"username": "cliente", "password": NEW_CLIENT_TEST_PASSWORD})
    logs = ok(client.get("/api/audit-logs?limit=20", headers=headers))["logs"]

    assert old_login.status_code == 401
    assert old_token.status_code == 401
    assert new_login.status_code == 200
    assert "user_password_changed" in {item["action"] for item in logs}


def test_self_password_change_requires_current_password_and_keeps_current_session(client):
    admin = login(client, "admin", ADMIN_TEST_PASSWORD)
    headers = auth_header(admin["token"])

    rejected = client.patch(
        "/api/users/admin/password",
        headers=headers,
        json={"password": NEW_ADMIN_TEST_PASSWORD},
    )
    changed = client.patch(
        "/api/users/admin/password",
        headers=headers,
        json={"currentPassword": ADMIN_TEST_PASSWORD, "password": NEW_ADMIN_TEST_PASSWORD},
    )
    current_token_still_valid = client.post("/api/auth/validate", json={"token": admin["token"]})
    new_login = client.post("/api/auth/login", json={"username": "admin", "password": NEW_ADMIN_TEST_PASSWORD})

    assert rejected.status_code == 403
    assert changed.status_code == 200
    assert current_token_still_valid.status_code == 200
    assert new_login.status_code == 200


def test_token_in_query_string_is_not_accepted(client):
    admin = login(client, "admin", ADMIN_TEST_PASSWORD)

    rejected = client.get(f"/api/records?token={admin['token']}")
    accepted = client.get("/api/records", headers=auth_header(admin["token"]))

    assert rejected.status_code == 401
    assert accepted.status_code == 200


def test_invalid_limit_parameters_return_400(client):
    admin = login(client, "admin", ADMIN_TEST_PASSWORD)
    headers = auth_header(admin["token"])

    records = client.get("/api/records?limit=abc", headers=headers)
    audit = client.get("/api/audit-logs?limit=abc", headers=headers)

    assert records.status_code == 400
    assert audit.status_code == 400
    assert "debe ser numerico" in records.get_json()["message"]
    assert "debe ser numerico" in audit.get_json()["message"]


def test_exported_records_escape_spreadsheet_formulas(client):
    admin = login(client, "admin", ADMIN_TEST_PASSWORD)
    user = login(client, "cliente", CLIENT_TEST_PASSWORD)
    headers = auth_header(admin["token"])

    ok(
        client.post(
            "/api/mobile",
            json={
                "action": "saveData",
                "token": user["token"],
                "driverName": "formula test",
                "plateNumber": "7777abc",
                "employeeCode": "1277",
                "ebap": "norte",
                "initialReading": 1,
                "finalReading": 3,
                "loadVolume": 2,
                "companyType": "particular",
                "companyName": "=HYPERLINK(\"https://evil.example\")",
                "characteristics": "@SUM(1,1)",
            },
        )
    )

    csv_response = client.get("/api/records/export.csv", headers=headers)
    rows = list(csv.reader(StringIO(csv_response.data.decode("utf-8"))))
    assert "Codigo funcionario" not in rows[0]
    assert rows[1][9].startswith("'=")
    assert rows[1][10].startswith("'@")

    xlsx_response = client.get("/api/records/export.xlsx", headers=headers)
    workbook = load_workbook(BytesIO(xlsx_response.data), read_only=True, data_only=False)
    sheet = workbook["Registros"]
    assert "Codigo funcionario" not in [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
    assert sheet["J2"].value.startswith("'=")
    assert sheet["K2"].value.startswith("'@")


def test_oversized_import_is_rejected(client):
    admin = login(client, "admin", ADMIN_TEST_PASSWORD)
    client.application.config["MAX_CONTENT_LENGTH"] = 128

    response = client.post(
        "/api/records/import",
        headers=auth_header(admin["token"]),
        data={"file": (BytesIO(b"x" * 256), "registros.csv")},
        content_type="multipart/form-data",
    )

    assert response.status_code == 413
    assert "Archivo demasiado grande" in response.get_json()["message"]
